import logging
import pandas as pd

from typing import List, Union

import openpyxl

logging.basicConfig(level=logging.DEBUG,
                    format='%(asctime)s - %(filename)s[line:%(lineno)5d] - %(levelname)s: %(message)s')
log = logging.getLogger('choas')


class Excel:
    def __init__(self):
        self.wb: openpyxl.Workbook = None

    def open(self, name: str):
        if not name:
            self.wb = openpyxl.Workbook()
        else:
            self.wb = openpyxl.load_workbook(name)

    def get_sheet_data(self, sheet_name: str):
        sheet = self.get_sheet(sheet_name)
        if not sheet:
            return list(), list()

        grid = list()
        for row in range(1, sheet.max_row + 1):
            grid.append(list())
            for col in range(1, sheet.max_column + 1):
                grid[row - 1].append(sheet.cell(row, col).value)
        return grid[0], grid[1:]

    def get_header(self, sheet_name: str):
        sheet = self.get_sheet(sheet_name)
        if not sheet:
            return list()
        header = list()
        for col in range(1, sheet.max_column + 1):
            header.append(sheet.cell(1, col).value)
        return header

    def get_row_data(self, sheet_name: str, row: int) -> list:
        sheet = self.get_sheet(sheet_name)
        if not sheet:
            return list()

        row_data = list()
        for col in range(1, sheet.max_column + 1):
            log.debug(sheet.cell(row, col).value)
            row_data.append(sheet.cell(row, col).value)
        return row_data

    def set_row_data(self, sheet_name: str, row: int, row_data: list):
        sheet = self.get_sheet(sheet_name)
        if not sheet:
            return None

        if len(row_data) != sheet.max_column:
            raise ValueError(
                f"row data: {len(row_data)} len must be current sheet max_column: {sheet.max_column}")

        for col in range(1, sheet.max_column + 1):
            sheet.cell(row, col).value = row_data[col - 1]

    def get_col_data(self, sheet_name: str, col: int):
        sheet = self.get_sheet(sheet_name)
        if not sheet:
            return list()

        col_data = list()
        for row in range(1, sheet.max_row + 1):
            col_data.append(sheet.cell(row, col).value)
        return col_data

    def set_col_data(self, sheet_name: str, col: int, col_data: list):
        sheet = self.get_sheet(sheet_name)
        if not sheet:
            return None

        if len(col_data) != sheet.max_row - 1:
            raise ValueError(
                f"col data len: {len(col_data)} must be current sheet max_row: {sheet.max_row - 1}")

        for row in range(2, sheet.max_row + 1):
            sheet.cell(row, col).value = col_data[row - 2]

    def get_cell_data(self, sheet_name: str, row: int, col: int):
        sheet = self.get_sheet(sheet_name)
        if not sheet:
            return None

        return sheet.cell(row, col).value

    def set_cell_data(self, sheet_name: str, row: int, col: int, value: any):
        sheet = self.get_sheet(sheet_name)
        if not sheet:
            return

        sheet.cell(row, col).value = value

    def get_cell(self, sheet_name: str, row: int, col: int):
        sheet = self.get_sheet(sheet_name)
        if not sheet:
            return None

        return sheet.cell(row, col).value

    def get_sheet(self, sheet_name: str):
        sheets = self.wb.worksheets
        for it in sheets:
            if it.title == sheet_name:
                return it
        return None

    def save(self, file_name: str):
        self.wb.save(file_name)
